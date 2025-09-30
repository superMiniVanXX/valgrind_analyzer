"""
Excel Report Generator for Valgrind Log Analyzer

This module provides functionality to generate comprehensive Excel reports
from analyzed Valgrind memory issues data.
"""

import os
from pathlib import Path
from typing import Optional, Dict, Any, List
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:
    raise ImportError(
        "openpyxl is required for Excel report generation. "
        "Install it with: pip install openpyxl"
    ) from e

from models import ClassifiedIssues, Statistics, MemoryIssue, IssueType, IssueSeverity


class ExcelReportError(Exception):
    """Custom exception for Excel report generation errors."""
    pass


class ExcelReporter:
    """
    Generates Excel reports from Valgrind memory analysis data.
    
    This class creates comprehensive Excel workbooks with multiple worksheets
    containing summary statistics, detailed issue listings, and analysis data.
    """
    
    def __init__(self):
        """Initialize the Excel reporter with default styling."""
        self.logger = logging.getLogger(__name__)
        
        # Define common styles
        self.header_font = Font(bold=True, size=12)
        self.title_font = Font(bold=True, size=14)
        self.critical_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        self.high_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
        self.header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.left_alignment = Alignment(horizontal="left", vertical="center")
        
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    def create_workbook(self) -> Workbook:
        """
        Create a new Excel workbook with basic setup.
        
        Returns:
            Workbook: A new openpyxl Workbook instance
            
        Raises:
            ExcelReportError: If workbook creation fails
        """
        try:
            workbook = Workbook()
            # Remove the default sheet
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])
            
            self.logger.info("Created new Excel workbook")
            return workbook
            
        except Exception as e:
            raise ExcelReportError(f"Failed to create Excel workbook: {str(e)}") from e
    
    def create_worksheet(self, workbook: Workbook, title: str) -> Worksheet:
        """
        Create a new worksheet in the workbook with the given title.
        
        Args:
            workbook: The Excel workbook to add the worksheet to
            title: The title/name for the new worksheet
            
        Returns:
            Worksheet: The newly created worksheet
            
        Raises:
            ExcelReportError: If worksheet creation fails
        """
        try:
            worksheet = workbook.create_sheet(title=title)
            self.logger.debug(f"Created worksheet: {title}")
            return worksheet
            
        except Exception as e:
            raise ExcelReportError(f"Failed to create worksheet '{title}': {str(e)}") from e
    
    def save_workbook(self, workbook: Workbook, output_path: str) -> None:
        """
        Save the workbook to the specified file path.
        
        Args:
            workbook: The Excel workbook to save
            output_path: The file path where to save the workbook
            
        Raises:
            ExcelReportError: If saving fails
        """
        try:
            # Ensure the directory exists
            output_dir = Path(output_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Save the workbook
            workbook.save(output_path)
            self.logger.info(f"Excel report saved to: {output_path}")
            
        except PermissionError as e:
            raise ExcelReportError(
                f"Permission denied when saving to '{output_path}'. "
                f"Make sure the file is not open in another application."
            ) from e
        except OSError as e:
            raise ExcelReportError(f"Failed to save Excel file to '{output_path}': {str(e)}") from e
        except Exception as e:
            raise ExcelReportError(f"Unexpected error saving Excel file: {str(e)}") from e
    
    def apply_header_style(self, worksheet: Worksheet, row: int, start_col: int = 1, end_col: Optional[int] = None) -> None:
        """
        Apply header styling to a row of cells.
        
        Args:
            worksheet: The worksheet to apply styling to
            row: The row number to style
            start_col: Starting column (default: 1)
            end_col: Ending column (if None, uses the last column with data)
        """
        if end_col is None:
            end_col = worksheet.max_column
        
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
    
    def auto_adjust_column_widths(self, worksheet: Worksheet) -> None:
        """
        Automatically adjust column widths based on content.
        
        Args:
            worksheet: The worksheet to adjust column widths for
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Set a reasonable width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary_sheet(self, workbook: Workbook, classified_issues: ClassifiedIssues) -> Worksheet:
        """
        Create a summary worksheet with overview statistics.
        
        Args:
            workbook: The Excel workbook to add the worksheet to
            classified_issues: The analyzed and classified memory issues data
            
        Returns:
            Worksheet: The created summary worksheet
            
        Raises:
            ExcelReportError: If summary sheet creation fails
        """
        try:
            worksheet = self.create_worksheet(workbook, "Summary")
            statistics = classified_issues.statistics
            
            # Title
            worksheet["A1"] = "Valgrind Memory Analysis Summary"
            worksheet["A1"].font = Font(bold=True, size=16)
            worksheet.merge_cells("A1:D1")
            worksheet["A1"].alignment = self.center_alignment
            
            # Overall Statistics Section
            row = 3
            worksheet[f"A{row}"] = "Overall Statistics"
            worksheet[f"A{row}"].font = self.title_font
            worksheet[f"A{row}"].fill = self.header_fill
            worksheet.merge_cells(f"A{row}:D{row}")
            
            row += 1
            worksheet[f"A{row}"] = "Total Issues:"
            worksheet[f"B{row}"] = statistics.total_issues
            worksheet[f"C{row}"] = "Total Bytes Lost:"
            worksheet[f"D{row}"] = statistics.total_bytes_lost
            
            row += 1
            worksheet[f"A{row}"] = "Total Blocks Lost:"
            worksheet[f"B{row}"] = statistics.total_blocks_lost
            
            # Critical Issues Highlighting
            critical_count = statistics.severity_distribution.get(IssueSeverity.CRITICAL, 0)
            high_count = statistics.severity_distribution.get(IssueSeverity.HIGH, 0)
            
            if critical_count > 0:
                row += 1
                worksheet[f"A{row}"] = "⚠️ Critical Issues:"
                worksheet[f"B{row}"] = critical_count
                worksheet[f"A{row}"].fill = self.critical_fill
                worksheet[f"B{row}"].fill = self.critical_fill
                worksheet[f"A{row}"].font = Font(bold=True)
                worksheet[f"B{row}"].font = Font(bold=True)
            
            if high_count > 0:
                row += 1
                worksheet[f"A{row}"] = "⚠️ High Priority Issues:"
                worksheet[f"B{row}"] = high_count
                worksheet[f"A{row}"].fill = self.high_fill
                worksheet[f"B{row}"].fill = self.high_fill
                worksheet[f"A{row}"].font = Font(bold=True)
                worksheet[f"B{row}"].font = Font(bold=True)
            
            # Issues by Type Section
            row += 3
            worksheet[f"A{row}"] = "Issues by Type"
            worksheet[f"A{row}"].font = self.title_font
            worksheet[f"A{row}"].fill = self.header_fill
            worksheet.merge_cells(f"A{row}:D{row}")
            
            row += 1
            # Headers
            headers = ["Issue Type", "Count", "Percentage", "Bytes Lost"]
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.thin_border
            
            # Data rows
            percentage_by_type = statistics.get_percentage_by_type()
            for issue_type in IssueType:
                if issue_type in statistics.issues_by_type and statistics.issues_by_type[issue_type] > 0:
                    row += 1
                    count = statistics.issues_by_type[issue_type]
                    percentage = percentage_by_type.get(issue_type, 0)
                    bytes_lost = statistics.bytes_by_type.get(issue_type, 0)
                    
                    # Format issue type name
                    type_name = issue_type.value.replace("_", " ").title()
                    
                    worksheet[f"A{row}"] = type_name
                    worksheet[f"B{row}"] = count
                    worksheet[f"C{row}"] = f"{percentage:.1f}%"
                    worksheet[f"D{row}"] = bytes_lost
                    
                    # Apply highlighting for critical issue types
                    if issue_type in [IssueType.DEFINITELY_LOST, IssueType.INVALID_READ, 
                                    IssueType.INVALID_WRITE, IssueType.USE_AFTER_FREE]:
                        for col in range(1, 5):
                            worksheet.cell(row=row, column=col).fill = self.critical_fill
                    elif issue_type == IssueType.POSSIBLY_LOST:
                        for col in range(1, 5):
                            worksheet.cell(row=row, column=col).fill = self.high_fill
                    
                    # Apply borders
                    for col in range(1, 5):
                        worksheet.cell(row=row, column=col).border = self.thin_border
            
            # Bytes Distribution Section
            row += 3
            worksheet[f"A{row}"] = "Memory Loss Distribution"
            worksheet[f"A{row}"].font = self.title_font
            worksheet[f"A{row}"].fill = self.header_fill
            worksheet.merge_cells(f"A{row}:D{row}")
            
            row += 1
            # Headers
            headers = ["Issue Type", "Bytes Lost", "Percentage", "Avg per Issue"]
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.thin_border
            
            # Data rows
            bytes_percentage_by_type = statistics.get_bytes_percentage_by_type()
            for issue_type in IssueType:
                if issue_type in statistics.bytes_by_type and statistics.bytes_by_type[issue_type] > 0:
                    row += 1
                    bytes_lost = statistics.bytes_by_type[issue_type]
                    percentage = bytes_percentage_by_type.get(issue_type, 0)
                    count = statistics.issues_by_type.get(issue_type, 1)
                    avg_per_issue = bytes_lost / count if count > 0 else 0
                    
                    # Format issue type name
                    type_name = issue_type.value.replace("_", " ").title()
                    
                    worksheet[f"A{row}"] = type_name
                    worksheet[f"B{row}"] = bytes_lost
                    worksheet[f"C{row}"] = f"{percentage:.1f}%"
                    worksheet[f"D{row}"] = f"{avg_per_issue:.1f}"
                    
                    # Apply highlighting for critical issue types
                    if issue_type in [IssueType.DEFINITELY_LOST, IssueType.INVALID_READ, 
                                    IssueType.INVALID_WRITE, IssueType.USE_AFTER_FREE]:
                        for col in range(1, 5):
                            worksheet.cell(row=row, column=col).fill = self.critical_fill
                    elif issue_type == IssueType.POSSIBLY_LOST:
                        for col in range(1, 5):
                            worksheet.cell(row=row, column=col).fill = self.high_fill
                    
                    # Apply borders
                    for col in range(1, 5):
                        worksheet.cell(row=row, column=col).border = self.thin_border
            
            # Top Issue Sources Section (if available)
            if statistics.top_sources:
                row += 3
                worksheet[f"A{row}"] = "Top Issue Sources"
                worksheet[f"A{row}"].font = self.title_font
                worksheet[f"A{row}"].fill = self.header_fill
                worksheet.merge_cells(f"A{row}:D{row}")
                
                row += 1
                worksheet[f"A{row}"] = "Source Location"
                worksheet[f"A{row}"].font = self.header_font
                worksheet[f"A{row}"].fill = self.header_fill
                worksheet[f"A{row}"].alignment = self.center_alignment
                worksheet[f"A{row}"].border = self.thin_border
                worksheet.merge_cells(f"A{row}:D{row}")
                
                for source in statistics.top_sources[:10]:  # Show top 10
                    row += 1
                    worksheet[f"A{row}"] = source
                    worksheet[f"A{row}"].border = self.thin_border
                    worksheet.merge_cells(f"A{row}:D{row}")
            
            # Auto-adjust column widths
            self.auto_adjust_column_widths(worksheet)
            
            self.logger.debug("Summary worksheet created successfully")
            return worksheet
            
        except Exception as e:
            raise ExcelReportError(f"Failed to create summary worksheet: {str(e)}") from e

    def create_detailed_sheet(self, workbook: Workbook, issues: List[MemoryIssue]) -> Worksheet:
        """
        Create a detailed worksheet with complete issue listings.
        
        Args:
            workbook: The Excel workbook to add the worksheet to
            issues: List of all memory issues to include in the detailed view
            
        Returns:
            Worksheet: The created detailed issues worksheet
            
        Raises:
            ExcelReportError: If detailed sheet creation fails
        """
        try:
            worksheet = self.create_worksheet(workbook, "Detailed Issues")
            
            # Title
            worksheet["A1"] = "Detailed Memory Issues Report"
            worksheet["A1"].font = Font(bold=True, size=16)
            worksheet.merge_cells("A1:H1")
            worksheet["A1"].alignment = self.center_alignment
            
            # Headers row
            row = 3
            headers = [
                "Issue Type",
                "Severity", 
                "Bytes",
                "Blocks",
                "Loss Record",
                "Primary Function",
                "Source Location",
                "Stack Trace"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.thin_border
            
            # Data rows
            row += 1
            for issue in issues:
                # Format issue type name
                issue_type_name = issue.issue_type.value.replace("_", " ").title()
                
                # Get primary function from stack trace (first frame with function name)
                primary_function = "Unknown"
                if issue.stack_trace:
                    for frame in issue.stack_trace:
                        if frame.function_name and frame.function_name != "???":
                            primary_function = frame.function_name
                            break
                
                # Get source location (from issue or first stack frame with source info)
                source_location = issue.source_location or "Unknown"
                if source_location == "Unknown" and issue.stack_trace:
                    for frame in issue.stack_trace:
                        if frame.source_file:
                            if frame.line_number:
                                source_location = f"{frame.source_file}:{frame.line_number}"
                            else:
                                source_location = frame.source_file
                            break
                
                # Format stack trace (first 3 frames for readability)
                stack_trace_text = "No stack trace"
                if issue.stack_trace:
                    trace_lines = []
                    for i, frame in enumerate(issue.stack_trace[:3]):  # Limit to first 3 frames
                        trace_lines.append(str(frame))
                    stack_trace_text = "\n".join(trace_lines)
                    if len(issue.stack_trace) > 3:
                        stack_trace_text += f"\n... ({len(issue.stack_trace) - 3} more frames)"
                
                # Populate row data
                worksheet.cell(row=row, column=1, value=issue_type_name)
                worksheet.cell(row=row, column=2, value=issue.severity.name)
                worksheet.cell(row=row, column=3, value=issue.bytes_count)
                worksheet.cell(row=row, column=4, value=issue.blocks_count)
                worksheet.cell(row=row, column=5, value=issue.loss_record)
                worksheet.cell(row=row, column=6, value=primary_function)
                worksheet.cell(row=row, column=7, value=source_location)
                worksheet.cell(row=row, column=8, value=stack_trace_text)
                
                # Apply severity-based highlighting
                fill_color = None
                if issue.severity == IssueSeverity.CRITICAL:
                    fill_color = self.critical_fill
                elif issue.severity == IssueSeverity.HIGH:
                    fill_color = self.high_fill
                
                # Apply formatting to all cells in the row
                for col in range(1, 9):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = self.thin_border
                    cell.alignment = self.left_alignment
                    
                    if fill_color:
                        cell.fill = fill_color
                    
                    # Special formatting for specific columns
                    if col in [3, 4]:  # Bytes and Blocks columns
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col == 8:  # Stack trace column - wrap text
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                # Set row height for stack trace readability
                worksheet.row_dimensions[row].height = max(60, len(issue.stack_trace) * 15)
                
                row += 1
            
            # Set column widths for better readability
            column_widths = {
                'A': 18,  # Issue Type
                'B': 12,  # Severity
                'C': 10,  # Bytes
                'D': 8,   # Blocks
                'E': 12,  # Loss Record
                'F': 25,  # Primary Function
                'G': 30,  # Source Location
                'H': 50   # Stack Trace
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # Add summary information at the top
            worksheet["A2"] = f"Total Issues: {len(issues)}"
            worksheet["A2"].font = Font(bold=True, size=11)
            
            # Add filters to the header row
            worksheet.auto_filter.ref = f"A{3}:H{row-1}"
            
            self.logger.debug(f"Detailed issues worksheet created with {len(issues)} issues")
            return worksheet
            
        except Exception as e:
            raise ExcelReportError(f"Failed to create detailed issues worksheet: {str(e)}") from e

    def create_issue_type_sheets(self, workbook: Workbook, classified_issues: ClassifiedIssues) -> None:
        """
        Create separate worksheets for each issue type.
        
        Args:
            workbook: The Excel workbook to add worksheets to
            classified_issues: The analyzed and classified memory issues data
        """
        try:
            # Define user-friendly sheet names for each issue type
            sheet_names = {
                IssueType.DEFINITELY_LOST: "Definitely Lost",
                IssueType.POSSIBLY_LOST: "Possibly Lost", 
                IssueType.STILL_REACHABLE: "Still Reachable",
                IssueType.INVALID_READ: "Invalid Reads",
                IssueType.INVALID_WRITE: "Invalid Writes",
                IssueType.USE_AFTER_FREE: "Use After Free",
                IssueType.OTHER: "Other Issues"
            }
            
            # Create a sheet for each issue type that has issues
            for issue_type, issues in classified_issues.issues_by_type.items():
                if issues:  # Only create sheet if there are issues of this type
                    sheet_name = sheet_names.get(issue_type, issue_type.value.replace("_", " ").title())
                    self.logger.debug(f"Creating sheet for {issue_type.value}: {len(issues)} issues")
                    
                    worksheet = self.create_worksheet(workbook, sheet_name)
                    self.populate_issue_type_sheet(worksheet, issues, issue_type)
                    
        except Exception as e:
            raise ExcelReportError(f"Failed to create issue type sheets: {str(e)}") from e
    
    def populate_issue_type_sheet(self, worksheet: Worksheet, issues: List[MemoryIssue], issue_type: IssueType) -> None:
        """
        Populate a worksheet with issues of a specific type.
        
        Args:
            worksheet: The worksheet to populate
            issues: List of memory issues of the specific type
            issue_type: The type of issues in this sheet
        """
        try:
            # Add title
            title = f"{issue_type.value.replace('_', ' ').title()} Issues ({len(issues)} total)"
            worksheet['A1'] = title
            worksheet['A1'].font = Font(bold=True, size=14)
            
            # Add summary info
            total_bytes = sum(issue.bytes_count for issue in issues)
            total_blocks = sum(issue.blocks_count for issue in issues)
            
            worksheet['A2'] = f"Total Bytes: {total_bytes:,}"
            worksheet['A3'] = f"Total Blocks: {total_blocks:,}"
            worksheet['A2'].font = Font(bold=True)
            worksheet['A3'].font = Font(bold=True)
            
            # Headers starting at row 5
            headers = [
                "Severity", "Bytes", "Blocks", "Loss Record", 
                "Primary Function", "Source Location", "Stack Trace"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Data rows starting at row 6
            for row, issue in enumerate(issues, 6):
                # Severity
                worksheet.cell(row=row, column=1, value=issue.severity.name)
                
                # Bytes and Blocks
                worksheet.cell(row=row, column=2, value=issue.bytes_count)
                worksheet.cell(row=row, column=3, value=issue.blocks_count)
                
                # Loss Record
                worksheet.cell(row=row, column=4, value=issue.loss_record)
                
                # Primary Function (first non-??? function in stack trace)
                primary_function = "Unknown"
                if issue.stack_trace:
                    for frame in issue.stack_trace:
                        if frame.function_name and frame.function_name != "???":
                            primary_function = frame.function_name
                            break
                worksheet.cell(row=row, column=5, value=primary_function)
                
                # Source Location
                source_location = issue.source_location or "Unknown"
                if source_location == "Unknown" and issue.stack_trace:
                    for frame in issue.stack_trace:
                        if frame.source_file:
                            if frame.line_number:
                                source_location = f"{frame.source_file}:{frame.line_number}"
                            else:
                                source_location = frame.source_file
                            break
                worksheet.cell(row=row, column=6, value=source_location)
                
                # Stack Trace (formatted)
                if issue.stack_trace:
                    stack_trace_text = "\n".join([
                        f"{i+1}. {frame}" for i, frame in enumerate(issue.stack_trace[:5])  # Limit to first 5 frames
                    ])
                    if len(issue.stack_trace) > 5:
                        stack_trace_text += f"\n... and {len(issue.stack_trace) - 5} more frames"
                else:
                    stack_trace_text = "No stack trace available"
                
                cell = worksheet.cell(row=row, column=7, value=stack_trace_text)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Apply styling and auto-adjust columns
            self.apply_header_style(worksheet, 5)
            self.auto_adjust_column_widths(worksheet)
            
            # Set row height for better readability
            for row in range(6, len(issues) + 6):
                worksheet.row_dimensions[row].height = 60  # Adjust for wrapped text
                
            self.logger.debug(f"Populated {issue_type.value} sheet with {len(issues)} issues")
            
        except Exception as e:
            raise ExcelReportError(f"Failed to populate {issue_type.value} sheet: {str(e)}") from e

    def create_statistics_sheet(self, workbook: Workbook, classified_issues: ClassifiedIssues) -> Worksheet:
        """
        Create a statistics worksheet with aggregated data and analysis.
        
        Args:
            workbook: The Excel workbook to add the worksheet to
            classified_issues: The analyzed and classified memory issues data
            
        Returns:
            Worksheet: The created statistics worksheet
        """
        # TODO: Implement statistics sheet creation
        worksheet = self.create_worksheet(workbook, "Statistics")
        return worksheet
    
    def generate_report(self, classified_issues: ClassifiedIssues, output_path: str) -> None:
        """
        Generate a complete Excel report from classified issues data.
        
        This is the main entry point for report generation. It creates a workbook
        and delegates to specific methods for each worksheet.
        
        Args:
            classified_issues: The analyzed and classified memory issues data
            output_path: Path where the Excel file should be saved
            
        Raises:
            ExcelReportError: If report generation fails
        """
        try:
            self.logger.info("Starting Excel report generation")
            
            # Validate input data
            if not classified_issues or not classified_issues.all_issues:
                raise ExcelReportError("No issues data provided for report generation")
            
            # Create workbook
            workbook = self.create_workbook()
            
            # Create summary worksheet
            self.create_summary_sheet(workbook, classified_issues)
            
            # Create separate sheets for each issue type
            self.create_issue_type_sheets(workbook, classified_issues)
            
            # Create statistics worksheet
            self.create_statistics_sheet(workbook, classified_issues)
            
            # Save the workbook
            self.save_workbook(workbook, output_path)
            
            self.logger.info("Excel report generation completed successfully")
            
        except ExcelReportError:
            raise
        except Exception as e:
            raise ExcelReportError(f"Unexpected error during report generation: {str(e)}") from e